import openpyxl
from faker import Faker
import random
import string
import secrets

def generate_excel_data():
    # Crear un objeto Faker
    fake = Faker()

    # Crear un libro de Excel y una hoja de trabajo
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Agregar encabezados
    sheet['A1'] = "nombre"
    sheet['B1'] = "apellidos"
    sheet['C1'] = "genero"
    sheet['D1'] = "provincia"
    sheet['E1'] = "municipio"

    sheet['F1'] = "usuario"
    sheet['G1'] = "correo"
    sheet['H1'] = "contraseña"
    
    sheet['I1'] = "grupo"
    sheet['J1'] = "año"
    sheet['K1'] = "rol_facultad"
    
    sheet['L1'] = "rol_universitario"
    sheet['M1'] = "solapin"
    
    sheet['N1'] = "carrera"
    sheet['O1'] = "carnet_identidad"
    sheet['P1'] = "id_expediente"
    
    list_user = []

    list_provincia = ["Guantanamo", "Santiago de Cuba", "Granma", "Holguin", "Las Tunas", 
                    "Camaguey", "Villa Clara", "Sancti Spíritus", "Ciego de Avila", "Cienfuegos",
                    "Matanzas", "Mayabeque", "Artemisa","La Habana","Pinar del Rio", "Isla de la Juventud"]
    
    list_group = ["IDF2101", "IDF2102", "IDF2103", "IDF2104", "IDF2105", 
                  "IDF2201", "IDF2202", "IDF2203", "IDF2204", "IDF2205", 
                  "IDF2301", "IDF2302", "IDF2303", "IDF2304", "IDF2305", 
                  "IDF2401", "IDF2402", "IDF2403", "IDF2404", "IDF2405"]
    
    list_rol = ["estudiante", "profesor guia", "profesor de año", "vice-decana/o"]
    
    list_years = [1, 2, 3, 4]
    
    list_group_copy = list_group.copy()

    dict_provincia_municipios = {
                    "Guantanamo":["Baracoa", "Caimanera", "El Salvador", "Guantánamo", "Imías", "Maisí", "Manuel Tames", "Niceto Pérez", "San Antonio del Sur", "Yateras"],
                    "Santiago de Cuba":["Contramaestre", "Guamá", "Mella", "Palma Soriano", "San Luis", "Santiago de Cuba", "Segundo Frente", "Niceto Pérez", "Songo-La Maya", "Tercer Frente"],
                    "Granma":["Bartolomé Masó", "Bayamo", "Buey Arriba", "Campechuela", "Cauto Cristo", "Guisa", "Jiguaní", "Manzanillo", "Media Luna", "Niquero", "Pilón", "Río Cauto", "Yara"],
                    "Holguin":["Antilla", "Báguanos", "Banes", "Cacocum", "Calixto García", "Cueto", "Frank País", "Gibara", "Holguín", "Mayarí", "Moa", "Rafael Freyre", "Sagua de Tánamo", "Urbano Noris"],
                    "Las Tunas":["Amancio", "Colombia", "Jesús Menéndez", "Jobabo", "Las Tunas", "Majibacoa", "Manatí", "Puerto Padre"],
                    "Camaguey":["Camagüey", "Carlos M. de Céspedes", "Esmeralda", "Florida", "Guáimaro", "Jimaguayú", "Minas", "Najasa", "Nuevitas", "Santa Cruz del Sur", "Sibanicú", "Sierra de Cubitas", "Vertientes"],
                    "Ciego de Avila":["Baraguá", "Bolivia", "Chambas", "Ciego de Ávila", "Ciro Redondo", "Florencia", "Majagua", "Morón", "Primero de Enero", "Venezuela"],
                    "Sancti Spíritus":["Cabaiguán", "Fomento", "Jatibonico", "La Sierpe", "Sancti Spíritus", "Taguasco", "Trinidad", "Yaguajay"],
                    "Villa Clara":["Caibarién", "Camajuaní", "Cifuentes", "Corralillo", "Encrucijada", "Manicaragua", "Placetas", "Quemado de Güines", "Ranchuelo", "San Juan de los Remedios", "Sagua la Grande", "Santa Clara", "Santo Domingo"],
                    "Cienfuegos":["Abreus", "Aguada de Pasajeros", "Cienfuegos", "Cruces", "Cumanayagua", "Lajas", "Palmira", "Rodas"],
                    "Matanzas":["Calimete", "Cárdenas", "Ciénaga de Zapata", "Colón", "Jagüey Grande", "Jovellanos", "Limonar", "Los Arabos", "Martí", "Matanzas", "Pedro Betancourt", "Perico", "Unión de Reyes"],
                    "Mayabeque":["Batabanó", "Bejucal", "Güines", "Jaruco", "Madruga", "Melena del Sur", "Nueva Paz", "Quivicán", "San José de las Lajas", "San Nicolás", "Santa Cruz del Norte"],
                    "La Habana":["Arroyo Naranjo", "Boyeros", "Centro Habana", "Cerro", "Cotorro", "Diez de Octubre", "Guanabacoa", "La Habana del Este", "La Habana Vieja", "La Lisa", "Marianao", "Playa", "Plaza de la Revolución", "Regla", "San Miguel del Padrón"],
                    "Artemisa":["Alquízar", "Artemisa", "Bahía Honda", "Bauta", "Caimito", "Candelaria", "Guanajay", "Güira de Melena", "Mariel", "San Antonio de los Baños", "San Cristóbal"],
                    "Pinar del Rio":["Consolación del Sur", "Guane", "La Palma", "Los Palacios", "Mantua", "Minas de Matahambre", "Pinar del Río", "San Juan y Martínez", "San Luis", "Sandino", "Viñales"],
                    "Isla de la Juventud":["Nueva Gerona"],
                        }

    row = 2  # Comenzar desde la segunda fila

    for prov in list_provincia:
        # Generar nombres y apellidos ficticios y agregarlos a la hoja de trabajo
        for i in range(150):  # Generar 50 filas de datos ficticios por provincia
            
            gender = random.choice(["Masculino", "Femenino"])
            
            municipio = random.choice(dict_provincia_municipios[prov])
            
            last_name = fake.last_name()
            last_name = last_name + " " + fake.last_name()
            
            name = fake.first_name()
            user = generate_user(list_user, name, last_name)
            
            email = f'{user}@uci.cu'
            password = 'mi55ing****'
            
            group = random.choice(list_group)
            year = group[4]
            group = str(group)
            year = int(year)
            
            carrera = 'Ingenieria en Ciencias Informaticas'
            
            solapin = random.randint(100000, 999999)
            carnet_identidad = random.randint(100000000, 999999999)
            
            id_expediente = random.randint(10000, 99999)
            id_expediente = f'EH{str(id_expediente)}'
            
            rol = random.choice(list_rol)
            if rol == "profesor guia":
                group = random.choice(list_group_copy)
                list_group_copy.remove(group)
                year = group[4]
                group = str(group)
                year = int(year)
                solapin = f'T{str(solapin)}'
                carnet_identidad = int(f'83{str(carnet_identidad)}')
                
                if len(list_group_copy) == 0:
                    list_rol.remove(rol)
            
            elif rol == "profesor de año":
                group=""
                year=random.choice(list_years)
                list_years.remove(year)
                solapin = f'T{str(solapin)}'
                carnet_identidad = int(f'83{str(carnet_identidad)}')
                if len(list_years) == 0:
                    list_rol.remove(rol)
                
            elif rol == "vice-decana/o":
                group=""
                year=""
                list_rol.remove(rol)
                rol_universitario = "vicedecano"
                solapin = f'T{str(solapin)}'
                carnet_identidad = int(f'83{str(carnet_identidad)}')
            
            else:
                rol_universitario = "estudiante"
                solapin = f'E{str(solapin)}'
                carnet_identidad = int(f'96{str(carnet_identidad)}')
                
            if 'profesor' in rol:
                rol_universitario = 'profesor'
                
            
            sheet.cell(row, 1, name)
            sheet.cell(row, 2, last_name)
            sheet.cell(row, 3, gender)
            sheet.cell(row, 4, prov)
            sheet.cell(row, 5, municipio)
            
            sheet.cell(row, 6, user)
            sheet.cell(row, 7, email)
            sheet.cell(row, 8, password)
            
            sheet.cell(row, 9, group)
            sheet.cell(row, 10, year)
            sheet.cell(row, 11, rol)
            
            sheet.cell(row, 12, rol_universitario)
            sheet.cell(row, 13, solapin)
            sheet.cell(row, 14, carrera)
            
            sheet.cell(row, 15, carnet_identidad)
            sheet.cell(row, 16, id_expediente)
            
            row += 1
            
        

    # Guardar el archivo de Excel
    workbook.save("user_ramdon.xlsx")
    
def generate_user(list_user, name, last_name):
    for pos, _ in enumerate(last_name):
        name = str(name).lower()
        if name not in list_user:
            list_user.append(name)
            return name
        else:
            user = f'{name}{last_name[:pos]}'.lower()
            if user not in list_user:
                list_user.append(user)
                return user

    
if __name__ == '__main__':
    generate_excel_data()
    # generate_user(list_user, name, last_name)
